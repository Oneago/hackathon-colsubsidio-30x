#!/usr/bin/env python3
"""Convierte 'BODEGAS Y STOCK.xlsx' a un unico JSON combinado (estilo respuesta de API).

Salida: inventory.json
Estructura:
{
  "meta": {...},
  "bodegas": [{ "id", "nombre" }],
  "stock": {
     "<slug_ubicacion>": {
        "nombre": "<titulo original de la hoja>",
        "items": [{ "nrArticulo", "articulo", "unidad", "stock" }]
     }
  }
}

Notas de mapeo:
- La columna 'SD' de las hojas de stock es la EXISTENCIA REAL -> campo `stock`.
- La columna 'CANTIDAD' es solo un indice secuencial de fila y se descarta.
- 'Nr.Articulo' puede venir vacio -> null.
"""
import json
import re
import unicodedata
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
XLSX = HERE / "BODEGAS Y STOCK.xlsx"
OUT = HERE / "inventory.json"

BODEGAS_SHEET = "BODEGAS DISPONIBLES"


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def clean_str(value):
    if value is None:
        return None
    s = str(value).replace("\xa0", " ").strip()
    return s or None


def clean_num(value):
    """Devuelve int cuando es entero exacto, si no float; None si vacio."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    s = str(value).strip().replace(",", ".")
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return s or None


def parse_bodegas(ws):
    bodegas = []
    for row in ws.iter_rows(values_only=True):
        # Formato: (None, indice, nombre)
        cells = [c for c in row]
        idx = next((c for c in cells if isinstance(c, (int, float))), None)
        nombre = clean_str(next((c for c in cells if isinstance(c, str)), None))
        if idx is None or nombre is None:
            continue
        if str(nombre).upper() == "BODEGAS":  # fila de encabezado
            continue
        bodegas.append({"id": int(idx), "nombre": nombre})
    return bodegas


def parse_stock(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [clean_str(c) for c in rows[0]]
    # Localiza columnas por nombre de encabezado (tolerante a variaciones)
    col = {}
    for i, h in enumerate(header):
        if not h:
            continue
        hu = unicodedata.normalize("NFKD", h).encode("ascii", "ignore").decode().upper()
        if hu.startswith("NR"):
            col["nr"] = i
        elif hu.startswith("ARTIC"):
            col["articulo"] = i
        elif hu.startswith("UNIDAD"):
            col["unidad"] = i
        elif hu == "SD":
            col["stock"] = i

    items = []
    for row in rows[1:]:
        articulo = clean_str(row[col["articulo"]]) if "articulo" in col else None
        if articulo is None:
            continue
        items.append({
            "nrArticulo": clean_num(row[col["nr"]]) if "nr" in col else None,
            "articulo": articulo,
            "unidad": clean_str(row[col["unidad"]]) if "unidad" in col else None,
            "stock": clean_num(row[col["stock"]]) if "stock" in col else None,
        })
    return items


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    bodegas = parse_bodegas(wb[BODEGAS_SHEET])

    stock = {}
    total_items = 0
    for ws in wb.worksheets:
        if ws.title == BODEGAS_SHEET:
            continue
        items = parse_stock(ws)
        stock[slugify(ws.title)] = {"nombre": ws.title.strip(), "items": items}
        total_items += len(items)

    payload = {
        "meta": {
            "fuente": XLSX.name,
            "totalBodegas": len(bodegas),
            "ubicacionesStock": len(stock),
            "totalItems": total_items,
        },
        "bodegas": bodegas,
        "stock": stock,
    }

    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Escrito {OUT} :: {len(bodegas)} bodegas, {len(stock)} ubicaciones, {total_items} items")


if __name__ == "__main__":
    main()
